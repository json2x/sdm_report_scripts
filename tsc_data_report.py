#!/usr/bin/env /home/jcm/env/sdmreports/bin/python

import os, glob, sys, re, time
from mydataframe import dataframe
from oracle import Oracle
from datetime import datetime, timedelta, date
import paramiko
import xlsxwriter
from openpyxl import load_workbook
from shutil import copyfile

APP_ROOT_DIR = '/home/jcm/projects/SDMReports'
OUTPUT_FILE_PATH = '/home/jcm/projects/SDMReports/OUTPUT_REPORT/TSC_Data_Report.xlsx'
#---------------------------
def main():
	
	dfObj = dataframe()
	
	cr_wo_handled = dataframe()
	cr_handled = dataframe()
	cr_wo_accepted = dataframe()
	tsc_wo_new = dataframe()
	tsc_wo_pending = dataframe()
	sdt_running = dataframe()
	
	userGrp = dataframe()
	oracle = Oracle()
	
	print('Connecting to SDM Database...', flush=True)
	oracle.connect('inoc_sdm_servicedesk_1', 'DB_Service_!23', '10.150.6.23', '1550', 'sdm')
	if oracle.connection:
		print('Connection Successful!', flush=True)
		
		today = date.today()
		yesterday = today - timedelta(1)
		
		#get sql query and store in variable
		print('Extracting CR_WO_Handled_DAR2 data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/CR_WO_Handled_DAR2.sql', 'r') as myfile:
			sql=myfile.read()
		sql = sql.format(startDate=yesterday.strftime("%Y-%m-%d %H:%M:%S"), endDate=today.strftime("%Y-%m-%d %H:%M:%S"))
		cr_wo_handled.createSQLResultDataframe(sql, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(cr_wo_handled.df.index)), flush=True)
		
		print('Extracting CR_Handled_DAR2 data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/CR_Handled_DAR2.sql', 'r') as myfile:
			sql=myfile.read()
		sql = sql.format(startDate=yesterday.strftime("%Y-%m-%d %H:%M:%S"), endDate=today.strftime("%Y-%m-%d %H:%M:%S"))
		cr_handled.createSQLResultDataframe(sql, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(cr_handled.df.index)), flush=True)
		
		print('Extracting CR_WO_Accepted_DAR2 data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/CR_WO_Accepted_DAR2.sql', 'r') as myfile:
			sql=myfile.read()
		sql = sql.format(startDate=yesterday.strftime("%Y-%m-%d %H:%M:%S"), endDate=today.strftime("%Y-%m-%d %H:%M:%S"))
		cr_wo_accepted.createSQLResultDataframe(sql, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(cr_wo_accepted.df.index)), flush=True)
		
		print('Extracting PDT-TT-SQDT_NEW_DAR2 data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/PDT-TT-SQDT_NEW_DAR2.sql', 'r') as myfile:
			sql=myfile.read()
		sql = sql.format(startDate=yesterday.strftime("%Y-%m-%d %H:%M:%S"), endDate=today.strftime("%Y-%m-%d %H:%M:%S"))
		tsc_wo_new.createSQLResultDataframe(sql, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(tsc_wo_new.df.index)), flush=True)
		
		print("")
		print('Extracting PENDING_TT-PDT-SDT-SQDT_WO data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/PENDING_TT-PDT-SDT-SQDT_WO.sql', 'r') as myfile:
			sql=myfile.read()
		tsc_wo_pending.createSQLResultDataframe(sql, oracle.db)
		#get_and_add_escalation_time('WO', tsc_wo_pending, oracle.db)
		escalationDf = query_escalation_time('WO', tsc_wo_pending, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(tsc_wo_pending.df.index)), flush=True)
		print('Pending WO Escalation Df: ({} rows returned.)'.format(len(escalationDf.df.index)), flush=True)
		print("")
		
		print('Extracting RUNNING_SDT data into dataframe...', end='', flush=True)
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/RUNNING_SDT.sql', 'r') as myfile:
			sql=myfile.read()
		sdt_running.createSQLResultDataframe(sql, oracle.db)
		get_and_add_escalation_time('SDT', sdt_running, oracle.db)
		#query_escalation_time('SDT', sdt_running, oracle.db)
		print('[OK] ({} rows returned.)'.format(len(sdt_running.df.index)), flush=True)
		
		print('')
		print('')
		print('Writing dataframes into an excel file.', flush=True)
		#book = load_workbook(OUTPUT_FILE_PATH)
		book = xlsxwriter.Workbook(OUTPUT_FILE_PATH)
		writer = dfObj.pd.ExcelWriter(OUTPUT_FILE_PATH, engine='xlsxwriter')
		writer.book = book
		
		print('cr_wo_handled to Sheet CR_WO_Handled_DAR2', end='', flush=True)
		cr_wo_handled.df.to_excel(writer, sheet_name='CR_WO_Handled_DAR2', index=False)
		print('[OK]', flush=True)
		
		print('cr_handled to Sheet CR_Handled_DAR2', end='', flush=True)
		cr_handled.df.to_excel(writer, sheet_name='CR_Handled_DAR2', index=False)
		print('[OK]', flush=True)
		
		print('cr_wo_accepted to Sheet CR_WO_Accepted_DAR2', end='', flush=True)
		cr_wo_accepted.df.to_excel(writer, sheet_name='CR_WO_Accepted_DAR2', index=False)
		print('[OK]', flush=True)
		
		print('tsc_wo to Sheet PDT_TT_SQDT_NEW_DAR2', end='', flush=True)
		tsc_wo_new.df.to_excel(writer, sheet_name='PDT_TT_SQDT_NEW_DAR2', index=False)
		print('[OK]', flush=True)
		
		print('tsc_wo to Sheet PENDING_TT_PDT_SDT_SQDT_WO', end='', flush=True)
		tsc_wo_pending.df.to_excel(writer, sheet_name='PENDING_TT_PDT_SDT_SQDT_WO', index=False)
		print('[OK]', flush=True)
		
		print('sdt_running to Sheet RUNNING_SDT', end='', flush=True)
		sdt_running.df.to_excel(writer, sheet_name='RUNNING_SDT', index=False)
		print('[OK]', flush=True)
		
		print('escalationDf to Sheet ESCALATION-TIME_TT_PDT_SDT_SQDT_WO', end='', flush=True)
		escalationDf.df.to_excel(writer, sheet_name='PENDING_WO_ESCALATION_TIME', index=False)
		print('[OK]', flush=True)
		
		#writer.save()
		writer.close()
		print('File saved to {}'.format(OUTPUT_FILE_PATH), flush=True)
		
		
		DUMP_FILEPATH = '/home/jcm/mnt/TX_FILES/TXREPORTS/TSC_Data_Report_' + time.strftime("%Y%m%d") + '.xlsx'
		print('Moving file to {}...'.format(DUMP_FILEPATH), flush=True, end='')
		copyfile(OUTPUT_FILE_PATH, DUMP_FILEPATH)
		print('[OK]', flush=True)
		
		#print('Writing DataFrame to CSV.')
		#ttwo.df.to_csv(APP_ROOT_DIR + '/output.csv', index=False)
		
		#Mount ip and directory below
		#10.150.20.104:/data/smbdir/TX/TXREPORTS
		#SendFile(OUTPUT_FILE_PATH)
			
		oracle.disconnect()
	else:
		print('Connection Failed!')
		
#--------------------------------------------
def get_and_add_escalation_time(type, dfObj, db):
	escalationDf = dataframe()
	if type == 'WO':
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/ESCALATION_TIME_TTWO.sql', 'r') as myfile:
			sql=myfile.read()
	elif type == 'SDT':
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/ESCALATION_TIME_SDT.sql', 'r') as myfile:
			sql=myfile.read()
			
	dfObj.df["EscalationTime"] = ""
	for i, row in dfObj.df.iterrows():
		query = sql.replace('{ticketid}', row['TicketID'])
		escalationDf.createSQLResultDataframe(query, db)
		if not escalationDf.df.empty:			
			dfObj.df.at[i, "EscalationTime"] = escalationDf.df.iloc[0, 1]
			
#--------------------------------------------
def query_escalation_time(type, dfObj, db):
	escalationDf = dataframe()
	if type == 'WO':
		with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/ESCALATION_TIME_TTWO.sql', 'r') as myfile:
			sql=myfile.read()
	#elif type == 'SDT':
	#	with open(APP_ROOT_DIR + '/SQL/TSC_REPORT_SQL/ESCALATION_TIME_SDT.sql', 'r') as myfile:
	#		sql=myfile.read()
	
	ticketIdStr = ""
	for i, row in dfObj.df.iterrows():
		ticketIdStr += "'{}'".format(row['TicketID'])
		if i < (len(dfObj.df.index)-1):
			ticketIdStr += ","
	
	query = sql.replace('{ticketid}', ticketIdStr)	
	escalationDf.createSQLResultDataframe(query, db)
	return escalationDf
		
	
#---------------------------
#--- --- ---
# Sending file via paramiko sftp
#--- --- ---
def SendFile(localFilename):
	remoteFilename = "TSC_Data_Report_" + time.strftime("%Y%m%d") + ".csv"
	SERVER_ADDRESS = ['10.103.23.239']
	SERVER_UNAME = 'dfuser'
	SERVER_PSWD = 'dfuser1'
	SERVER_DIR = '/home/dfuser' 
	#'SDM_REPORTS/PENDING_TT-WOs'
	print("")
	print('Preparing to send file to Proj USA Datahub dump directory.')
	print('     {}'.format(localFilename))
	
	for ServerIP in SERVER_ADDRESS:
		try:
			print('Connecting to host: {} ... '.format(ServerIP), end="")
			t = paramiko.Transport((ServerIP, 22))
			t.connect(username=SERVER_UNAME, password=SERVER_PSWD)
			sftp = paramiko.SFTPClient.from_transport(t)
			sftp.chdir(SERVER_DIR)
			print('[OK]')
			
			print('Transferring {} to server... '.format(localFilename), end="")
			sftp.put(localFilename, remoteFilename)
			print('[OK]')
			
			print("")
			print("-- FILE SENT ---")
			print("")
			
			sftp.close()
		except Exception as e:
			print('[FAIL]')
			print('*** Caught exception: %s: %s' % (e.__class__, e))
				
		finally:
			t.close()

#---------------------------
if __name__ == "__main__":
	starTime = datetime.now()
	print("\nRun Datetime: {}".format(starTime))
	print("-------------------------")
	print("SDM Reports")
	print("TSC-Transport Data Report")
	print("-------------------------")
	
	main()
	
	endTime = datetime.now()
	deltaDt = endTime - starTime
	print(deltaDt)
	print("<<<<< End of Script >>>>>")
