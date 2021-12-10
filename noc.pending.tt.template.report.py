#!/usr/bin/env /home/jcm/env/sdmreports/bin/python

import os, glob, sys, re, time
from mydataframe import dataframe
from oracle import Oracle
from datetime import datetime, timedelta
import paramiko
import xlsxwriter
from openpyxl import load_workbook

APP_ROOT_DIR = '/home/jcm/projects/SDMReports'
OUTPUT_FILE_PATH = ''
#---------------------------
def main():
	
	tt = dataframe()
	oracle = Oracle()
	
	try:
		print('Connecting to SDM Database...', end="", flush=True)
		oracle.connect('inoc_sdm_servicedesk_1', 'DB_Service_!23', '10.150.6.23', '1550', 'sdm')
		if oracle.connection:
			print('[OK]')
			print('Connection Successful!')
			print('')
			#get sql query and store in variable
			print('Querying and Creating TT Dataframe...', end="", flush=True)
			with open(APP_ROOT_DIR + '/SQL/pending.tt.sql', 'r') as myfile:
				sql=myfile.read()
			tt.createSQLResultDataframe(sql, oracle.db)
			print('[OK]')
			
			print('')
			print('Cleansing and Prepping Trouble Ticket dataframe')
			tt = PrepAndCleanseData(tt)
			print('')
			print(tt.df)
			
			LuzViMinTT = tt.df[(tt.df['Area'] == 'MM') | (tt.df['Area'] == 'NL') | (tt.df['Area'] == 'SL') \
			| (tt.df['Area'] == 'VIS') | (tt.df['Area'] == 'MIN')]
			LuzViMinxTech = LuzViMinTT.pivot_table(index='Area', columns='Technology', values='TicketID', margins=True, fill_value="", aggfunc='count')
			
			print(LuzViMinxTech)
			
			tech_list = ['2G','3G','FD-LTE','5G','Canopy','TD-LTE','WIFI']
			area_list = ['MM','NL','SL','VIS','MIN']
			
			
			OUTPUT_FILE_PATH = '/home/jcm/projects/SDMReports/REPORT_TEMPLATE/AUTO_TT_REPORT.xlsx'
			workbook = xlsxwriter.Workbook(OUTPUT_FILE_PATH)
			
			#FORMATS
			tbl_measure_cell_format = workbook.add_format({'align': 'center', 'valign': 'middle', 'border': 1})
			tbl_header_cell_format = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'bg_color': '#C7DDFC', 'border': 1})
			tbl_dimension_cell_format = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'border': 1})
			tbl_grand_total_measure_cell_format = workbook.add_format({'align': 'center', 'valign': 'middle', \
			'bg_color': '#C7DDFC', 'border': 1})
			tbl_grand_total_dimension_cell_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': \
			'middle', 'bg_color': '#C7DDFC', 'border': 1})
			
			
			#-------- Sheet Preparation ---------
			print('*** Preparing 1st Sheet ***')
			worksheet1 = workbook.add_worksheet()
			cell_format = workbook.add_format({'bold': True, 'font_size': '20'})
			worksheet1.write('A1', 'OUTAGE TROUBLE TICKETS', cell_format)
			worksheet1.set_column(0, 0, 23)
			cell_format = workbook.add_format({'bold': True, 'font_size': '14'})
			worksheet1.write('A3', 'Per Technology', cell_format)
			
			worksheet1.write('A4', 'Area', tbl_header_cell_format)
			ctr = 1
			for tech in tech_list:
				worksheet1.write(3, ctr, tech, tbl_header_cell_format)
				ctr += 1
			worksheet1.write(3, ctr, 'TOTAL', tbl_header_cell_format)
			
			#----- Populating data in Sheet ----
			print('Creating reshaped dataframes.')
			
			vpos = 4
			for area in area_list:
				worksheet1.write(vpos, 0, area, tbl_dimension_cell_format)
				writeMultiColInTable(worksheet1, LuzViMinxTech, area, tech_list, vpos, 1, True, tbl_measure_cell_format)
				vpos += 1
			worksheet1.write(vpos, 0, 'TOTAL', tbl_grand_total_dimension_cell_format)
			writeMultiColInTable(worksheet1, LuzViMinxTech, 'All', tech_list, vpos, 1, True, tbl_grand_total_measure_cell_format)
			print('')
			
			workbook.close()
			oracle.disconnect()
			
			print('Including Raw Data to output file...', end="", flush=True)
			book = load_workbook(OUTPUT_FILE_PATH)
			writer = tt.pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl')
			writer.book = book
			tt.df.to_excel(writer, sheet_name='Raw Data', index=False)
			#uniqueSiteIDDF.to_excel(writer, sheet_name='UniqSID_SA Raw Data', index=False)
			writer.save()
			writer.close()
			print('[OK]')
			print('')
			
			print('Sending file via FTP...')
			SendFile(OUTPUT_FILE_PATH)
			
		else:
			print('Connection Failed!')
	except Exception as e:
		print('[FAILED]')
		print('*** Unexpected error: %s: %s ***' % (e.__class__, e))

#---------------------------------------
def writeSingleColInTable(worksheet, sourceDf, index_lbl, col_lbl, vpos, hpos, cell_format = None):
	try:
		if index_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index_lbl]]
			if cell_format:
				worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col_lbl], cell_format)
			else:
				worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col_lbl])
		else:
			writeBlankCell(worksheet, vpos, hpos, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------
def writeMultiColInTable(worksheet, sourceDf, index_lbl, col_list, vpos, hpos, rowTotal = False, cell_format = None):
	try:
		if index_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index_lbl]]
			for col in col_list:
				if col in sliceDf:
					if cell_format:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col], cell_format)
					else:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col])
				else:
					writeBlankCell(worksheet, vpos, hpos, cell_format)
				hpos += 1
			if rowTotal:
				if 'All' in sliceDf:
					if cell_format:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, 'All'], cell_format)
					else:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, 'All'])
		else:
			writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))

#---------------------------------------
def writeMultiIndexSingleColInTable(worksheet, sourceDf, index1_lbl, index2_lbl, col_lbl, vpos, hpos, cell_format = None):
	try:
		if index1_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index1_lbl]]
			for iTuple, row in sliceDf.iterrows():
				if index2_lbl in iTuple and col_lbl in row:
					if cell_format:
						worksheet.write(vpos, hpos, row[col_lbl], cell_format)
					else:
						worksheet.write(vpos, hpos, row[col_lbl])
					break
		else:
			writeBlankCell(worksheet, vpos, hpos, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------
def writeMultiIndexMultiColInTable(worksheet, sourceDf, index1_lbl, index2_lbl, col_list, vpos, hpos, rowTotal = False, cell_format = None):
	try:
		if index1_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index1_lbl]]
			for iTuple, row in sliceDf.iterrows():
				if index2_lbl in iTuple:
					for col in col_list:
						if col in row:
							if cell_format:
								worksheet.write(vpos, hpos, row[col], cell_format)
							else:
								worksheet.write(vpos, hpos, row[col])
						else:
							writeBlankCell(worksheet, vpos, hpos, cell_format)
						hpos += 1
						
					if rowTotal:
						if 'All' in sliceDf:
							worksheet.write(vpos, hpos, row['All'])
					break
				else:
					writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
					
		else:
			writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
			
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------		
def writeBlankCell(worksheet, vpos, hpos, cell_format):
	if cell_format:
		worksheet.write(vpos, hpos, '', cell_format)
	else:
		worksheet.write(vpos, hpos, '')

#---------------------------------------
def writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format):
	for col in col_list:
		if cell_format:
			worksheet.write(vpos, hpos, '', cell_format)
		else:
			worksheet.write(vpos, hpos, '')
		hpos += 1
	if rowTotal:
		if cell_format:
			worksheet.write(vpos, hpos, '', cell_format)
		else:
			worksheet.write(vpos, hpos, '')

#---------------------------------------
def PrepAndCleanseData(tt):
	
	print("Prepping/Cleansing TT Dataframe...", end="", flush=True)
	if not tt.df.empty:
				
		tt.df['Area'] = tt.df['Area'].replace('Metro Manila', 'MM')
		tt.df['Area'] = tt.df['Area'].replace('North Luzon', 'NL')
		tt.df['Area'] = tt.df['Area'].replace('South Luzon', 'SL')
		tt.df['Area'] = tt.df['Area'].replace('Visayas', 'VIS')
		tt.df['Area'] = tt.df['Area'].replace('Mindanao', 'MIN')
		
		tt.df['Technology'] = tt.df['Technology'].replace('FDLTE', 'FD-LTE')
		tt.df['Technology'] = tt.df['Technology'].replace('TDLTE', 'TD-LTE')
		
		#tt.df['Fault_Severity_Level'] = tt.pd.Categorical(tt.df['Fault_Severity_Level'], ['Critical', 'Moderate', 'Major', 'Minor', 'Low', 'Alarm Clearing'])
		
		print('[Complete]')
		return tt
		
	else:
		print('Dataframe is empty. Can not proceed with processing.')
		print('Script will terminate.')
		
		return None
		
#---------------------------------------
def computeDuration(dfRow):
	duration = None
	
	deltaDt = datetime.today() - datetime.strptime(dfRow['WO_CreateTime'] ,"%Y-%m-%d %H:%M:%S")
	duration = strfdelta(deltaDt, "{days} {hours} {minutes}")
	duration_arr = duration.split(" ")
	duration_hrs = (int(duration_arr[0])*24) + int(duration_arr[1]) + (int(duration_arr[2])/60)
		
	return duration_hrs

#---------------------------------------	
def getAgingGroup(duration):
	agingGrp = None
	if duration < 4:
		agingGrp = '<4Hrs'
	elif 4 <= duration <= 24:
		agingGrp = '4Hrs-24Hrs'
	elif duration > 24:
		agingGrp = '>24Hrs'
		
	return agingGrp

#---------------------------------------	
def getAreaInProcessorGroup(currentProcessorGroup):
	newArea = None
	searchPatternList = {
		'Metro Manila': re.compile('^NFS_[A-Z]*MM_[A-Z0-9]*$'),
		'North Luzon': re.compile('^NFS_[A-Z]*NL_[A-Z0-9]*$'),
		'South Luzon': re.compile('^NFS_[A-Z]*SL_[A-Z0-9]*$'),
		'Visayas': re.compile('^NFS_[A-Z]*VIS_[A-Z0-9]*$'),
		'Mindanao': re.compile('^NFS_[A-Z]*MIN_[A-Z0-9]*$')
	}
	
	for area, regex in searchPatternList.items():
		test = regex.search(currentProcessorGroup)
		if test:
			newArea = area
			break
			
	return newArea
	
#---------------------------------------
def identifyWorkgroup(currentProcessor_group, currentProcessor):
	workGroup = None
	searchPhrase = currentProcessor_group
	if searchPhrase == 'NA':
		searchPhrase = currentProcessor
		
	#Search Patterns
	searchPatternList = {
		'WFS': re.compile(r'WFS'),
		'FFS': re.compile(r'FFS|FxATOP'),
		'TSC': re.compile(r'TSC|TAC'),
		'BUILD': re.compile(r'Build'),
		'NPE': re.compile(r'Engineering|NPE'),
		'NOSM': re.compile(r'SFEM'),
		'SPM': re.compile(r'SPMONS'),
		'FCFFS': re.compile(r'FC'),
		'CoreOps': re.compile(r'CoreOps'),
		'CSOG': re.compile(r'CSOG'),
		'ECSA': re.compile(r'ECSA'),
		'FNAS': re.compile(r'FNAS'),
		'FSMG': re.compile(r'FSMG'),
		'OPM': re.compile(r'OPM'),
		'OVIM': re.compile(r'OVIM'),
		'PMO': re.compile(r'PMO'),
		'NOC': re.compile(r'NOC'),
		'SOC': re.compile(r'SOC')
	}
	
	for group, regex in searchPatternList.items():
		test = regex.search(searchPhrase)
		if test:
			workGroup = group
			break
			
	return workGroup
	
#---------------------------------------		
def get_tt_orderid(ticketid, db):
	#If TT_TicketID is empty, check query back to db its TT_TicketID(Orderid) using the parentticketid field.
	cur = db.cursor()
	cur.execute('SELECT orderid FROM tbl_troubleticket_datasource WHERE ticketid = {}'.format(ticketid))
	results = cur.fetchall()
	TT_TicketID = ''
	for result in results:
		if not TT_TicketID:
			TT_TicketID = result[0]
		else:
			TT_TicketID = ';{}'.format(result[0])
	
	return TT_TicketID

#---------------------------------------	
def get_user_defaultgroup(user, userGrpDf):
	resultDf = userGrpDf[(userGrpDf['FULLNAME']==user)]
	group = 'NA'
	if not resultDf.empty:
		for i, row in resultDf.iterrows():
			group = row['user_default_group']
			break
	else:
		resultDf = userGrpDf[(userGrpDf['user_default_group']==user)]
		if not resultDf.empty:
			group = user
			
	return group
	
#---------------------------------------
def strfdelta(deltaDt, format):
	d = {"days": deltaDt.days}
	d['hours'], rem = divmod(deltaDt.seconds, 3600)
	d['minutes'], d['seconds'] = divmod(rem, 60)
	
	d['hours'] = '{:02d}'.format(d['hours'])
	d['minutes'] = '{:02d}'.format(d['minutes'])
	d['seconds'] = '{:02d}'.format(d['seconds'])
	
	return format.format(**d)
	
#--- --- ---
# Sending file via paramiko sftp
#--- --- ---
def SendFile(localFilename):
	remoteFilename = "AUTO_TT_REPORT_" + time.strftime("%Y%m%d.%H%M") + ".xlsx"
	SERVER_ADDRESS = ['10.150.20.104']
	SERVER_UNAME = 'smbnoc'
	SERVER_PSWD = 'smbnoc'
	SERVER_DIR = 'PENDING_TT-WOs' 
	#'SDM_REPORTS/PENDING_TT-WOs'
	print("")
	print('Preparing to send file to Proj USA Datahub dump directory.')
	print('     {}'.format(localFilename))
	
	for ServerIP in SERVER_ADDRESS:
		try:
			print('Connecting to host: {} ... '.format(ServerIP), end="")
			t = paramiko.Transport((ServerIP, 22))
			t.connect(username=SERVER_UNAME, password=SERVER_PSWD)
			sftp = paramiko.SFTPClient.from_transport(t)
			sftp.chdir(SERVER_DIR)
			print('[OK]')
			
			print('Transferring {} to server... '.format(localFilename), end="")
			sftp.put(localFilename, remoteFilename)
			print('[OK]')
			
			print("")
			print("-- FILE SENT ---")
			print("")
			
			sftp.close()
		except Exception as e:
			print('[FAIL]')
			print('*** Caught exception: %s: %s' % (e.__class__, e))
				
		finally:
			t.close()

#---------------------------
if __name__ == "__main__":
	starTime = datetime.now()
	print("\nRun Datetime: {}".format(starTime))
	print("-------------------------")
	print("SDM Reports")
	print("Pending/Running TroubleTickets")
	print("-------------------------")
	
	main()
	
	endTime = datetime.now()
	deltaDt = endTime - starTime
	print(deltaDt)
	print("<<<<< End of Script >>>>>")
