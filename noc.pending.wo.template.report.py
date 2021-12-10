#!/usr/bin/env /home/jcm/env/sdmreports/bin/python

import os, glob, sys, re, time
from mydataframe import dataframe
from oracle import Oracle
from datetime import datetime, timedelta
import paramiko
import xlsxwriter
from openpyxl import load_workbook

APP_ROOT_DIR = '/home/jcm/projects/SDMReports'
OUTPUT_FILE_PATH = ''
#---------------------------
def main():
	
	ttwo = dataframe()
	userGrp = dataframe()
	oracle = Oracle()
	currYear = datetime.now().year
	try:
		print('Connecting to SDM Database...', end="", flush=True)
		oracle.connect('inoc_sdm_servicedesk_1', 'DB_Service_!23', '10.150.6.23', '1550', 'sdm')
		if oracle.connection:
			print('[OK]')
			print('Connection Successful!')
			print('')
			#get sql query and store in variable
			print('Querying and Creating TT-WO Dataframe...', end="", flush=True)
			with open(APP_ROOT_DIR + '/SQL/pending.tt.wos.v2.sql', 'r') as myfile:
				sql=myfile.read()
			ttwo.createSQLResultDataframe(sql, oracle.db)
			print('[OK]')
			
			print('Querying and Creating User Default Group Dataframe...', end="", flush=True)
			query = 'SELECT user_id, fullname, get_default_gorup(user_id) as "user_default_group" from v_um_user where active = 1 and get_default_gorup(user_id) is not null'
			userGrp.createSQLResultDataframe(query, oracle.db)
			print('[OK]')
			#print(userGrp.df)
			
			print('')
			print('Cleansing and Prepping TT-WO dataframe')
			ttwo = PrepAndCleanseData(ttwo, userGrp)
			print('')
			
			'''
			print('Writing DataFrame to CSV...', end="", flush=True)
			ttwo.df.to_csv(APP_ROOT_DIR + '/auto_ttwo_report.csv', index=False)
			#SendFile(APP_ROOT_DIR + '/output.csv')
			print('[OK]')
			'''
			
			tech_list = ['2G','3G','FD-LTE','5G','Canopy','TD-LTE','WIFI']
			area_list = ['MM','NL','SL','VIS','MIN']
			agingGrp_list = ['<4Hrs','4Hrs-24Hrs','>24Hrs']
			sat_list = ['SA', 'PSA']
			
			workgroup_list = ttwo.df.Workgroup.unique()
			workgroup_list = workgroup_list.tolist()
			workgroup_list = [i for i in workgroup_list if i] #Remove None from list
			workgroup_list.remove('NOC') #Remove NOC group from workgroup list as requested by Requestor
			workgroup_list.sort()
			
			
			businessStatus_list = ttwo.df.WO_BusinessStatus.unique()
			businessStatus_list = businessStatus_list.tolist()
			businessStatus_list = [i for i in businessStatus_list if i] #Remove None from list
			businessStatus_list.sort()
			
			
			OUTPUT_FILE_PATH = '/home/jcm/projects/SDMReports/REPORT_TEMPLATE/AUTO_TTWO_REPORT.xlsx'
			workbook = xlsxwriter.Workbook(OUTPUT_FILE_PATH)
			#-------- Sheet 1 Preparation ---------
			print('*** Preparing 1st Sheet ***')
			worksheet1 = workbook.add_worksheet()
			cell_format = workbook.add_format({'bold': True, 'font_size': '20'})
			worksheet1.write('A1', 'BTS OUTAGE', cell_format)
			worksheet1.set_column(0, 0, 23)
			cell_format = workbook.add_format({'bold': True, 'font_size': '14'})
			worksheet1.write('A3', 'Per Technology', cell_format)
			tbl_header_cell_format = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'bg_color': '#C7DDFC', 'border': 1})
			#tbl_header_cell_format_red_font = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			#'valign': 'middle', 'bg_color': '#C7DDFC', 'border': 1, 'font_color': 'red'})
			tbl_header_cell_format_yellow = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'bg_color': '#EAF29D', 'border': 1})
			tbl_header_cell_format_red = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'bg_color': '#F2B8B8', 'border': 1})
			tbl_header_cell_format_grey = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'bg_color': '#BFBFBF', 'border': 1})
			tbl_dimension_cell_format = workbook.add_format({'bold': True, 'text_wrap': True, 'align': 'center', \
			'valign': 'middle', 'border': 1})
			tbl_measure_cell_format = workbook.add_format({'align': 'center', 'valign': 'middle', 'border': 1})
			tbl_measure_cell_format_ave_aging = workbook.add_format({'align': 'center', 'valign': 'middle', 'border': 1, 'num_format': '#,###,##0.00'})
			tbl_grand_total_dimension_cell_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': \
			'middle', 'bg_color': '#C7DDFC', 'border': 1})
			tbl_grand_total_measure_cell_format = workbook.add_format({'align': 'center', 'valign': 'middle', \
			'bg_color': '#C7DDFC', 'border': 1})
			tbl_grand_total_measure_cell_format_ave_aging = workbook.add_format({'align': 'center', 'valign': 'middle', \
			'bg_color': '#C7DDFC', 'border': 1, 'num_format': '#,###,##0.00'})
			worksheet1.write('A4', 'Area', tbl_header_cell_format)
			ctr = 1
			for tech in tech_list:
				worksheet1.write(3, ctr, tech, tbl_header_cell_format)
				ctr += 1
			worksheet1.write(3, ctr, 'TOTAL', tbl_header_cell_format)
			
			worksheet1.write('A12', 'Per Resolving Team', cell_format)
			worksheet1.merge_range('A13:A15', 'WorkGroup', tbl_header_cell_format)
			worksheet1.merge_range('B13:F13', 'AGEING (UNIQUE SITES)', tbl_header_cell_format)
			worksheet1.merge_range('B14:B15', 'Average (Hrs)', tbl_header_cell_format_yellow)
			worksheet1.merge_range('C14:C15', 'Prev Years', tbl_header_cell_format)
			worksheet1.merge_range('D14:F14', currYear, tbl_header_cell_format)
			ctr = 3
			for agingGrp in agingGrp_list:
				worksheet1.write(14, ctr, agingGrp, tbl_header_cell_format)
				ctr += 1
			
			#----- Populating data in Sheet1 ----
			
			print('Creating reshaped dataframes.')
			#***
			#print('Outage TTWO Pivot Table (Area x Technology)...', end="", flush=True)
			serviceAffectingDF = ttwo.df[ttwo.df['SiteDown'] == 'isYes']
			areaXtechDF = serviceAffectingDF.pivot_table(index='Area', columns='Technology', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			#print('[OK]')
			
			print(areaXtechDF)
			vpos = 4
			for area in area_list:
				worksheet1.write(vpos, 0, area, tbl_dimension_cell_format)
				#writeAreaPerTechRow(worksheet1, areaXtechDF, area, tech_list, vpos)
				writeMultiColInTable(worksheet1, areaXtechDF, area, tech_list, vpos, 1, True, tbl_measure_cell_format)
				vpos += 1
			worksheet1.write(vpos, 0, 'TOTAL', tbl_grand_total_dimension_cell_format)
			#writeAreaPerTechRow(worksheet1, areaXtechDF, 'All', tech_list, vpos)
			writeMultiColInTable(worksheet1, areaXtechDF, 'All', tech_list, vpos, 1, True, tbl_grand_total_measure_cell_format)
			print('')
			
			print('Unique SiteID DF')
			uniqueSiteIDAveAging = serviceAffectingDF.groupby('SiteID')['Aging (Hrs)'].mean().reset_index()
			uniqueSiteIDAveAging.dropna(subset=['Aging (Hrs)'], inplace=True)
			
			uniqueSiteIDDF = serviceAffectingDF.sort_values("Fault_Severity_Level")
			uniqueSiteIDDF.dropna(subset=['SiteID'], inplace=True)
			uniqueSiteIDDF.drop_duplicates(subset=['SiteID'], inplace=True, ignore_index=True)
			uniqueSiteIDDF.drop(columns='Aging (Hrs)', inplace=True)
			
			uniqueSiteIDDF = ttwo.pd.merge(uniqueSiteIDDF, uniqueSiteIDAveAging, how='left', on='SiteID').reset_index()
			print(uniqueSiteIDDF)
			print('')
			
			#workgroupAveAgingDF = serviceAffectingDF.pivot_table(index='Workgroup', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			workgroupAveAgingDF = uniqueSiteIDDF.pivot_table(index='Workgroup', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			print(workgroupAveAgingDF)
			print('')
			
			#prevYearServiceAffectingDF = ttwo.df[(ttwo.df['SiteDown'] == 'isYes') & (ttwo.df['WO_CreateYear'] < currYear)]
			#currYearServiceAffectingDF = ttwo.df[(ttwo.df['SiteDown'] == 'isYes') & (ttwo.df['WO_CreateYear'] == currYear)]
			prevYearServiceAffectingDF = uniqueSiteIDDF[uniqueSiteIDDF['WO_CreateYear'] < currYear]
			currYearServiceAffectingDF = uniqueSiteIDDF[uniqueSiteIDDF['WO_CreateYear'] == currYear]
			workgroupXPrevYearDF = prevYearServiceAffectingDF.pivot_table(index='Workgroup', columns='WO_Status', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			workgroupXAgingGrpDF = currYearServiceAffectingDF.pivot_table(index='Workgroup', columns='Aging Grp', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			print(workgroupXPrevYearDF)
			print('')
			print(workgroupXAgingGrpDF)
			print('')
			
			vpos = 15
			for workgroup in workgroup_list:
				worksheet1.write(vpos, 0, workgroup, tbl_dimension_cell_format)
				writeSingleColInTable(worksheet1, workgroupAveAgingDF, workgroup, 'Aging (Hrs)', vpos, 1, tbl_measure_cell_format_ave_aging)
				writeSingleColInTable(worksheet1, workgroupXPrevYearDF, workgroup, 'running', vpos, 2, tbl_measure_cell_format)
				writeMultiColInTable(worksheet1, workgroupXAgingGrpDF, workgroup, agingGrp_list, vpos, 3, False, tbl_measure_cell_format)
				vpos += 1
			worksheet1.write(vpos, 0, 'TOTAL', tbl_grand_total_dimension_cell_format)
			writeSingleColInTable(worksheet1, workgroupAveAgingDF, 'All', 'Aging (Hrs)', vpos, 1, tbl_grand_total_measure_cell_format_ave_aging)
			writeSingleColInTable(worksheet1, workgroupXPrevYearDF, 'All', 'running', vpos, 2, tbl_grand_total_measure_cell_format)
			writeMultiColInTable(worksheet1, workgroupXAgingGrpDF, 'All', agingGrp_list, vpos, 3, False, tbl_grand_total_measure_cell_format)
			print('*** 1st Sheet Complete ***')
			print('')
			
			#-------- Sheet 2 Preparation ---------
			print('*** Preparing 2nd Sheet ***')
			worksheet2 = workbook.add_worksheet()
			worksheet2.set_column(0, 0, 23)
			cell_format = workbook.add_format({'bold': True, 'font_size': '16'})
			worksheet2.write('A2', 'SUMMARY OF SDM TT-WO', cell_format)
			worksheet2.set_row(2, 30)
			worksheet2.set_row(3, 30)
			worksheet2.merge_range('A3:A4', 'SEVERITY', tbl_header_cell_format)
			worksheet2.merge_range('B3:B4', 'TOTAL PENDING', tbl_header_cell_format)
			worksheet2.merge_range('C3:D3', 'With FRT', tbl_header_cell_format)
			worksheet2.merge_range('E3:H3', 'AGEING', tbl_header_cell_format)
			worksheet2.write('C4', 'COUNT', tbl_header_cell_format)
			worksheet2.write('D4', 'Ageing (Ave. Hrs)', tbl_header_cell_format)
			worksheet2.write('E4', 'Ave Aging (Hours) Now', tbl_header_cell_format)
			ctr = 5
			for agingGrp in agingGrp_list:
				worksheet2.write(3, ctr, agingGrp, tbl_header_cell_format)
				ctr += 1
			worksheet2.merge_range(2, 8, 2, 8+(len(businessStatus_list)-1), 'BUSINESS STATUS', tbl_header_cell_format)
			ctr = 8
			for businessStatus in businessStatus_list:
				if businessStatus == 'Resovled':
					businessStatus = 'Resolved'
				worksheet2.write(3, ctr, businessStatus, tbl_header_cell_format)
				ctr += 1
				
			cell_format = workbook.add_format({'bold': True, 'font_size': '16'})
			worksheet2.write('A10', 'SDM TT-WO per Workgroup', cell_format)
			worksheet2.set_row(10, 30)
			worksheet2.set_row(11, 30)
			worksheet2.merge_range('A11:A12', 'WORKGROUP', tbl_header_cell_format)
			worksheet2.merge_range('B11:B12', 'SEVERITY', tbl_header_cell_format)
			worksheet2.merge_range('C11:C12', 'TOTAL', tbl_header_cell_format)
			worksheet2.merge_range('D11:E11', 'With FRT', tbl_header_cell_format_red)
			worksheet2.write('D12', 'COUNT', tbl_header_cell_format)
			worksheet2.write('E12', 'AGEING (AVE)', tbl_header_cell_format)
			worksheet2.merge_range('F11:I11', 'AGEING', tbl_header_cell_format_yellow)
			worksheet2.write('F12', 'AVE (HRS)', tbl_header_cell_format)
			ctr = 6
			for agingGrp in agingGrp_list:
				worksheet2.write(11, ctr, agingGrp, tbl_header_cell_format)
				ctr += 1
				
			worksheet2.merge_range(10, 9, 10, 9+(len(businessStatus_list)-1), 'BUSINESS STATUS', tbl_header_cell_format_grey)
			ctr = 9
			for businessStatus in businessStatus_list:
				if businessStatus == 'Resovled':
					businessStatus = 'Resolved'
				worksheet2.write(11, ctr, businessStatus, tbl_header_cell_format)
				ctr += 1
			#businessStatus_list
			
			print('Creating reshaped dataframes.')
			print('All Pending data')
			totalPendingDF = ttwo.df.pivot_table(index='Service Affecting Type', columns='WO_Status', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			aveAgingDF = ttwo.df.pivot_table(index='Service Affecting Type', columns='WO_Status', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			agingGrpDF = ttwo.df.pivot_table(index='Service Affecting Type', columns='Aging Grp', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			businessStatusDF = ttwo.df.pivot_table(index='Service Affecting Type', columns='WO_BusinessStatus', values='WO_Status', margins=True, fill_value="", aggfunc='count')
			print(totalPendingDF)
			print('')
			print(aveAgingDF)
			print('')
			print(agingGrpDF)
			print('')
			print(businessStatusDF)
			print('')
			
			print('With FRT data')
			withFRTDF = ttwo.df[ttwo.df.TT_FaultRecoveryTime.notnull()]
			SAvPSAwithFRTCountDF = withFRTDF.pivot_table(index='Service Affecting Type', columns='WO_Status', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			print(SAvPSAwithFRTCountDF)
			print('')
			SAvPSAwithFRTAveAgingDF = withFRTDF.pivot_table(index='Service Affecting Type', columns='WO_Status', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			print(SAvPSAwithFRTAveAgingDF)
			print('')
			
			vpos = 4
			for sat in sat_list:
				worksheet2.write(vpos, 0, sat, tbl_dimension_cell_format)
				writeSingleColInTable(worksheet2, totalPendingDF, sat, 'running', vpos, 1, tbl_measure_cell_format)
				writeSingleColInTable(worksheet2, SAvPSAwithFRTCountDF, sat, 'running', vpos, 2, tbl_measure_cell_format)
				writeSingleColInTable(worksheet2, SAvPSAwithFRTAveAgingDF, sat, 'running', vpos, 3, tbl_measure_cell_format_ave_aging)
				writeSingleColInTable(worksheet2, aveAgingDF, sat, 'running', vpos, 4, tbl_measure_cell_format_ave_aging)
				writeMultiColInTable(worksheet2, agingGrpDF, sat, agingGrp_list, vpos, 5, False, tbl_measure_cell_format)
				writeMultiColInTable(worksheet2, businessStatusDF, sat, businessStatus_list, vpos, 8, False, tbl_measure_cell_format)
				vpos += 1
			worksheet2.write(vpos, 0, 'All', tbl_grand_total_dimension_cell_format)
			writeSingleColInTable(worksheet2, totalPendingDF, 'All', 'running', vpos, 1, tbl_grand_total_dimension_cell_format)
			writeSingleColInTable(worksheet2, SAvPSAwithFRTCountDF, 'All', 'running', vpos, 2, tbl_grand_total_measure_cell_format)
			writeSingleColInTable(worksheet2, SAvPSAwithFRTAveAgingDF, 'All', 'running', vpos, 3, tbl_grand_total_measure_cell_format_ave_aging)
			writeSingleColInTable(worksheet2, aveAgingDF, 'All', 'running', vpos, 4, tbl_grand_total_measure_cell_format_ave_aging)
			writeMultiColInTable(worksheet2, agingGrpDF, 'All', agingGrp_list, vpos, 5, False, tbl_grand_total_measure_cell_format)
			writeMultiColInTable(worksheet2, businessStatusDF, 'All', businessStatus_list, vpos, 8, False, tbl_grand_total_measure_cell_format)
			
			
			print('')
			print('----------------------')
			print('')
			
			totalPendingPerWorkgroupPerSeverity = ttwo.df.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='WO_Status', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			
			aveAgingPerWorkgroupPerSeverity = ttwo.df.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='WO_Status', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			
			agingGrpPerWorkgroupPerSeverity = ttwo.df.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='Aging Grp', values='WO_Status', margins=True, fill_value="", aggfunc='count')
			
			businessStatusPerWorkgroupPerSeverity = ttwo.df.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='WO_BusinessStatus', values='WO_Status', margins=True, fill_value="", aggfunc='count')
			
			withFRTPendingPerWorkgroupPerSeverity = withFRTDF.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='WO_Status', values='WO_TicketID', margins=True, fill_value="", aggfunc='count')
			
			withFRTaveAgingPerWorkgroupPerSeverity = withFRTDF.pivot_table(index=['Workgroup','Service Affecting Type'], \
			columns='WO_Status', values='Aging (Hrs)', margins=True, fill_value="", aggfunc='mean')
			
			print(totalPendingPerWorkgroupPerSeverity)
			print(aveAgingPerWorkgroupPerSeverity)
			print(agingGrpPerWorkgroupPerSeverity)
			print(businessStatusPerWorkgroupPerSeverity)
			print(withFRTPendingPerWorkgroupPerSeverity)
			print(withFRTaveAgingPerWorkgroupPerSeverity)
			
			vpos = 12
			for workgroup in workgroup_list:
				worksheet2.merge_range(vpos, 0, vpos+1, 0, workgroup, tbl_dimension_cell_format)
				for sat in sat_list:
					worksheet2.write(vpos, 1, sat, tbl_dimension_cell_format)
					writeMultiIndexSingleColInTable(worksheet2, totalPendingPerWorkgroupPerSeverity, workgroup, sat, 'running', vpos, 2, tbl_measure_cell_format)
					writeMultiIndexSingleColInTable(worksheet2, withFRTPendingPerWorkgroupPerSeverity, workgroup, sat, 'running', vpos, 3, tbl_measure_cell_format)
					writeMultiIndexSingleColInTable(worksheet2, withFRTaveAgingPerWorkgroupPerSeverity, workgroup, sat, 'running', vpos, 4, tbl_measure_cell_format_ave_aging)
					writeMultiIndexSingleColInTable(worksheet2, aveAgingPerWorkgroupPerSeverity, workgroup, sat, 'running', vpos, 5, tbl_measure_cell_format_ave_aging)
					writeMultiIndexMultiColInTable(worksheet2, agingGrpPerWorkgroupPerSeverity, workgroup, sat, agingGrp_list, vpos, 6, False, tbl_measure_cell_format)
					writeMultiIndexMultiColInTable(worksheet2, businessStatusPerWorkgroupPerSeverity, workgroup, sat, businessStatus_list, vpos, 9 , False, tbl_measure_cell_format)
					vpos += 1
			
			print('[Complete]')
			workbook.close()
			oracle.disconnect()
			
			print('Including Raw Data to output file...', end="", flush=True)
			book = load_workbook(OUTPUT_FILE_PATH)
			writer = ttwo.pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl')
			writer.book = book
			ttwo.df.to_excel(writer, sheet_name='All Raw Data', index=False)
			#uniqueSiteIDDF.to_excel(writer, sheet_name='UniqSID_SA Raw Data', index=False)
			writer.save()
			writer.close()
			print('[OK]')
			print('')
			
			print('Sending file via FTP...')
			SendFile(OUTPUT_FILE_PATH)
			
		else:
			print('Connection Failed!')
	except Exception as e:
		print('[FAILED]')
		print('*** Unexpected error: %s: %s ***' % (e.__class__, e))

#---------------------------------------
def writeSingleColInTable(worksheet, sourceDf, index_lbl, col_lbl, vpos, hpos, cell_format = None):
	try:
		if index_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index_lbl]]
			if cell_format:
				worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col_lbl], cell_format)
			else:
				worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col_lbl])
		else:
			writeBlankCell(worksheet, vpos, hpos, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------
def writeMultiColInTable(worksheet, sourceDf, index_lbl, col_list, vpos, hpos, rowTotal = False, cell_format = None):
	try:
		if index_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index_lbl]]
			for col in col_list:
				if col in sliceDf:
					if cell_format:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col], cell_format)
					else:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, col])
				else:
					writeBlankCell(worksheet, vpos, hpos, cell_format)
				hpos += 1
			if rowTotal:
				if 'All' in sliceDf:
					if cell_format:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, 'All'], cell_format)
					else:
						worksheet.write(vpos, hpos, sourceDf.loc[index_lbl, 'All'])
		else:
			writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))

#---------------------------------------
def writeMultiIndexSingleColInTable(worksheet, sourceDf, index1_lbl, index2_lbl, col_lbl, vpos, hpos, cell_format = None):
	try:
		if index1_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index1_lbl]]
			for iTuple, row in sliceDf.iterrows():
				if index2_lbl in iTuple and col_lbl in row:
					if cell_format:
						worksheet.write(vpos, hpos, row[col_lbl], cell_format)
					else:
						worksheet.write(vpos, hpos, row[col_lbl])
					break
		else:
			writeBlankCell(worksheet, vpos, hpos, cell_format)
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------
def writeMultiIndexMultiColInTable(worksheet, sourceDf, index1_lbl, index2_lbl, col_list, vpos, hpos, rowTotal = False, cell_format = None):
	try:
		if index1_lbl in sourceDf.index:
			sliceDf = sourceDf.loc[[index1_lbl]]
			for iTuple, row in sliceDf.iterrows():
				if index2_lbl in iTuple:
					for col in col_list:
						if col in row:
							if cell_format:
								worksheet.write(vpos, hpos, row[col], cell_format)
							else:
								worksheet.write(vpos, hpos, row[col])
						else:
							writeBlankCell(worksheet, vpos, hpos, cell_format)
						hpos += 1
						
					if rowTotal:
						if 'All' in sliceDf:
							worksheet.write(vpos, hpos, row['All'])
					break
				else:
					writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
					
		else:
			writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format)
			
	except Exception as e:
		print('* Function error: %s: %s *' % (e.__class__, e))
		
#---------------------------------------		
def writeBlankCell(worksheet, vpos, hpos, cell_format):
	if cell_format:
		worksheet.write(vpos, hpos, '', cell_format)
	else:
		worksheet.write(vpos, hpos, '')

#---------------------------------------
def writeBlankRowCells(worksheet, col_list, vpos, hpos, rowTotal, cell_format):
	for col in col_list:
		if cell_format:
			worksheet.write(vpos, hpos, '', cell_format)
		else:
			worksheet.write(vpos, hpos, '')
		hpos += 1
	if rowTotal:
		if cell_format:
			worksheet.write(vpos, hpos, '', cell_format)
		else:
			worksheet.write(vpos, hpos, '')

#---------------------------------------
def PrepAndCleanseData(ttwo, userGrp):
	#Insert a workgroup column
	ttwo.df.insert(11,"Workgroup", "")
	ttwo.df.insert(8,"Aging Grp", "")
	ttwo.df.insert(22,"Service Affecting Type", "")
	
	print("Iterating through TT-WO Dataframe...", end="", flush=True)
	if not ttwo.df.empty and not userGrp.df.empty:
		for i, row in ttwo.df.iterrows():
			agingDuration = computeDuration(row)
			ttwo.df.at[i, "Aging (Hrs)"] = agingDuration
			ttwo.df.at[i, "Aging Grp"] = getAgingGroup(agingDuration)
			if ttwo.pd.isnull(row['TT_TicketID']):
				#If TT_TicketID is empty, check query back to db its TT_TicketID(Orderid) using the parentticketid field.
				ttwo.df.at[i, "TT_TicketID"] = get_tt_orderid(row['PARENTTICKETID'], oracle.db)
				
			if row['Technology'] == 'FDLTE':
				ttwo.df.at[i, "Technology"] = 'FD-LTE'
			elif row['Technology'] == 'TDLTE':
				ttwo.df.at[i, "Technology"] = 'TD-LTE'
				
			current_processor = str(row['WO_CurrentProcessor'])
			if current_processor.find(';') !=- 1:
				proccessors = row['WO_CurrentProcessor'].split(';')
				current_processor_group = get_user_defaultgroup_from_list(proccessors, userGrp.df)
				#current_processor_group = get_user_defaultgroup(proccessors[0], userGrp.df)
			else:
				current_processor_group = get_user_defaultgroup(row['WO_CurrentProcessor'], userGrp.df)
				 
			ttwo.df.at[i, "WO_CurrentProcessorGroup"] = current_processor_group
			ttwo.df.at[i, "Workgroup"] = identifyWorkgroup(current_processor_group, current_processor)
			
			if row['Area'] != 'Metro Manila' and row['Area'] != 'North Luzon' and row['Area'] != 'South Luzon' \
			and row['Area'] != 'Visayas' and row['Area'] != 'Mindanao':
				wo_title = str(row['WO_Title'])
				wo_title_arr = wo_title.split('|')
				areaInTitle = wo_title_arr[0].strip()
				if areaInTitle != 'Metro Manila' and areaInTitle != 'North Luzon' and  areaInTitle != 'South Luzon' \
				and areaInTitle != 'Visayas' and areaInTitle != 'Mindanao':
					newArea = getAreaInProcessorGroup(current_processor_group)
					ttwo.df.at[i, "Area"] = newArea
				else:
					ttwo.df.at[i, "Area"] = areaInTitle
					
			if row['SiteDown'] == 'isNo':
				ttwo.df.at[i, "Service Affecting Type"] = 'PSA'
			elif row['SiteDown'] == 'isYes':
				ttwo.df.at[i, "Service Affecting Type"] = 'SA'
			else:
				ttwo.df.at[i, "Service Affecting Type"] = None
				
		ttwo.df = ttwo.df.drop(columns=['PARENTTICKETID'])
		ttwo.df['Area'] = ttwo.df['Area'].replace('Metro Manila', 'MM')
		ttwo.df['Area'] = ttwo.df['Area'].replace('North Luzon', 'NL')
		ttwo.df['Area'] = ttwo.df['Area'].replace('South Luzon', 'SL')
		ttwo.df['Area'] = ttwo.df['Area'].replace('Visayas', 'VIS')
		ttwo.df['Area'] = ttwo.df['Area'].replace('Mindanao', 'MIN')
		ttwo.df['Aging (Hrs)'] = ttwo.pd.to_numeric(ttwo.df['Aging (Hrs)'])
		ttwo.df['WO_CreateYear'] = ttwo.pd.to_numeric(ttwo.df['WO_CreateYear'])
		
		ttwo.df['Fault_Severity_Level'] = ttwo.pd.Categorical(ttwo.df['Fault_Severity_Level'], ['Critical', 'Moderate', 'Major', 'Minor', 'Low', 'Alarm Clearing'])
		
		print('[Complete]')
		return ttwo
		
	else:
		print('Either or both dataframe is empty. Can not proceed with processing.')
		print('Script will terminate.')
		
		return None
		
#---------------------------------------
def computeDuration(dfRow):
	duration = None
	
	deltaDt = datetime.today() - datetime.strptime(dfRow['WO_CreateTime'] ,"%Y-%m-%d %H:%M:%S")
	duration = strfdelta(deltaDt, "{days} {hours} {minutes}")
	duration_arr = duration.split(" ")
	duration_hrs = (int(duration_arr[0])*24) + int(duration_arr[1]) + (int(duration_arr[2])/60)
		
	return duration_hrs

#---------------------------------------	
def getAgingGroup(duration):
	agingGrp = None
	if duration < 4:
		agingGrp = '<4Hrs'
	elif 4 <= duration <= 24:
		agingGrp = '4Hrs-24Hrs'
	elif duration > 24:
		agingGrp = '>24Hrs'
		
	return agingGrp

#---------------------------------------	
def getAreaInProcessorGroup(currentProcessorGroup):
	newArea = None
	searchPatternList = {
		'Metro Manila': re.compile('^NFS_[A-Z]*MM_[A-Z0-9]*$'),
		'North Luzon': re.compile('^NFS_[A-Z]*NL_[A-Z0-9]*$'),
		'South Luzon': re.compile('^NFS_[A-Z]*SL_[A-Z0-9]*$'),
		'Visayas': re.compile('^NFS_[A-Z]*VIS_[A-Z0-9]*$'),
		'Mindanao': re.compile('^NFS_[A-Z]*MIN_[A-Z0-9]*$')
	}
	
	for area, regex in searchPatternList.items():
		test = regex.search(currentProcessorGroup)
		if test:
			newArea = area
			break
			
	return newArea
	
#---------------------------------------
def identifyWorkgroup(currentProcessor_group, currentProcessor):
	workGroup = None
	searchPhrase = currentProcessor_group
	if searchPhrase == 'NA':
		searchPhrase = currentProcessor
		
	#Search Patterns
	searchPatternList = {
		'WFS': re.compile(r'WFS'),
		'FFS': re.compile(r'FFS|FxATOP'),
		'TSC': re.compile(r'TSC|TAC'),
		'BUILD': re.compile(r'Build'),
		'NPE': re.compile(r'Engineering|NPE'),
		'NOSM': re.compile(r'SFEM'),
		'SPM': re.compile(r'SPMONS'),
		'FCFFS': re.compile(r'FC'),
		'CoreOps': re.compile(r'CoreOps'),
		'CSOG': re.compile(r'CSOG'),
		'ECSA': re.compile(r'ECSA'),
		'FNAS': re.compile(r'FNAS'),
		'FSMG': re.compile(r'FSMG'),
		'OPM': re.compile(r'OPM'),
		'OVIM': re.compile(r'OVIM'),
		'PMO': re.compile(r'PMO'),
		'NOC_BIS': re.compile(r'NOC_BIS'),
		'NOC_CNCC': re.compile(r'NOC_CNCC'),
		'NOC': re.compile(r'NOC_NAC_NSC'),
		'SOC': re.compile(r'SOC')
	}
	
	for group, regex in searchPatternList.items():
		test = regex.search(searchPhrase)
		if test:
			workGroup = group
			break
			
	return workGroup
	
#---------------------------------------		
def get_tt_orderid(ticketid, db):
	#If TT_TicketID is empty, check query back to db its TT_TicketID(Orderid) using the parentticketid field.
	cur = db.cursor()
	cur.execute('SELECT orderid FROM tbl_troubleticket_datasource WHERE ticketid = {}'.format(ticketid))
	results = cur.fetchall()
	TT_TicketID = ''
	for result in results:
		if not TT_TicketID:
			TT_TicketID = result[0]
		else:
			TT_TicketID = ';{}'.format(result[0])
	
	return TT_TicketID

#---------------------------------------	
def get_user_defaultgroup(user, userGrpDf):
	resultDf = userGrpDf[(userGrpDf['FULLNAME']==user)]
	group = 'NA'
	if not resultDf.empty:
		for i, row in resultDf.iterrows():
			group = row['user_default_group']
			break
	else:
		resultDf = userGrpDf[(userGrpDf['user_default_group']==user)]
		if not resultDf.empty:
			group = user
			
	return group
	
#---------------------------------------
def get_user_defaultgroup_from_list(currentProcessorList, userGrpDf):
	group = 'NA'
	match_startswith_list = ['Engineering', 'NPE', 'Network Build', 'TSC', 'TAC', 'SFEM', 'OPM', 'WFS', 'NOC_BIS', \
	'FNAS', 'FFS', 'FxATOP', 'CNCC', 'SPMONS', 'FC', 'CoreOps', 'CSOG', 'OPM' 'OVIM', 'PMO', 'SOC', 'NOC']
	for grpCode in match_startswith_list:
		result = match_group_code_in_list(grpCode, currentProcessorList)
		if result:
			group = result[0]
			break
		
	if group == 'NA':
		group = get_user_defaultgroup(currentProcessorList[0], userGrpDf)
		
	return group

#---------------------------------------
def match_group_code_in_list(searchStr, itemList):

	return list(filter(lambda item: searchStr in item, itemList))
	
#---------------------------------------
def strfdelta(deltaDt, format):
	d = {"days": deltaDt.days}
	d['hours'], rem = divmod(deltaDt.seconds, 3600)
	d['minutes'], d['seconds'] = divmod(rem, 60)
	
	d['hours'] = '{:02d}'.format(d['hours'])
	d['minutes'] = '{:02d}'.format(d['minutes'])
	d['seconds'] = '{:02d}'.format(d['seconds'])
	
	return format.format(**d)
	
#--- --- ---
# Sending file via paramiko sftp
#--- --- ---
def SendFile(localFilename):
	remoteFilename = "AUTO_TT-WO_REPORT_" + time.strftime("%Y%m%d.%H%M") + ".xlsx"
	SERVER_ADDRESS = ['10.150.20.104']
	SERVER_UNAME = 'smbnoc'
	SERVER_PSWD = 'smbnoc'
	SERVER_DIR = 'PENDING_TT-WOs' 
	#'SDM_REPORTS/PENDING_TT-WOs'
	print("")
	print('Preparing to send file to Proj USA Datahub dump directory.')
	print('     {}'.format(localFilename))
	
	for ServerIP in SERVER_ADDRESS:
		try:
			print('Connecting to host: {} ... '.format(ServerIP), end="")
			t = paramiko.Transport((ServerIP, 22))
			t.connect(username=SERVER_UNAME, password=SERVER_PSWD)
			sftp = paramiko.SFTPClient.from_transport(t)
			sftp.chdir(SERVER_DIR)
			print('[OK]')
			
			print('Transferring {} to server... '.format(localFilename), end="")
			sftp.put(localFilename, remoteFilename)
			print('[OK]')
			
			print("")
			print("-- FILE SENT ---")
			print("")
			
			sftp.close()
		except Exception as e:
			print('[FAIL]')
			print('*** Caught exception: %s: %s' % (e.__class__, e))
				
		finally:
			t.close()

#---------------------------
if __name__ == "__main__":
	starTime = datetime.now()
	print("\nRun Datetime: {}".format(starTime))
	print("-------------------------")
	print("SDM Reports")
	print("Pending/Running TT-WOs")
	print("-------------------------")
	
	main()
	
	endTime = datetime.now()
	deltaDt = endTime - starTime
	print(deltaDt)
	print("<<<<< End of Script >>>>>")
